"""Cash Status Router - Kassenstand-Übersicht"""
import logging
from datetime import date, datetime
from io import BytesIO, StringIO
from typing import Optional
import csv

from fastapi import APIRouter, Request, Depends, Query
from fastapi.responses import HTMLResponse, Response
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func, union_all, literal
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.pdfgen import canvas as pdf_canvas

from app.database import get_db
from app.models import Payment, Expense, Income, Participant, Family, Event, Ruleset, Role
from app.dependencies import get_current_event_id
from app.templates_config import templates
from app.services.excel_service import ExcelService
from app.services.price_calculator import PriceCalculator

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/cash-status", tags=["cash_status"])


def calculate_base_prices_sum(db: Session, event_id: int) -> float:
    """
    Berechnet die Summe der Basispreise (ohne Rabatte) für alle Teilnehmer eines Events.

    Args:
        db: Datenbank-Session
        event_id: ID des Events

    Returns:
        Summe der Basispreise in Euro
    """
    # Event laden
    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        return 0.0

    # Aktives Ruleset für das Event finden
    ruleset = db.query(Ruleset).filter(
        Ruleset.event_id == event_id,
        Ruleset.is_active == True,
        Ruleset.valid_from <= event.start_date,
        Ruleset.valid_until >= event.start_date
    ).first()

    if not ruleset:
        return 0.0

    # Alle aktiven Teilnehmer laden
    participants = db.query(Participant).filter(
        Participant.event_id == event_id,
        Participant.is_active == True
    ).all()

    total_base_price = 0.0

    for participant in participants:
        # Alter berechnen
        age = event.start_date.year - participant.birth_date.year
        if (event.start_date.month, event.start_date.day) < (participant.birth_date.month, participant.birth_date.day):
            age -= 1

        # Basispreis aus Altersgruppen ermitteln (ohne Rabatte)
        base_price = PriceCalculator._get_base_price_by_age(
            age,
            ruleset.age_groups or []
        )

        total_base_price += base_price

    return round(total_base_price, 2)


def calculate_non_subsidy_discount_sum(db: Session, event_id: int) -> float:
    """
    Berechnet die Summe der Rabatte für nicht-zuschussberechtigte Rollen.
    Diese Rabatte werden auf die Gruppe umgelegt und nicht durch Zuschüsse gedeckt.

    Args:
        db: Datenbank-Session
        event_id: ID des Events

    Returns:
        Summe der nicht-zuschussberechtigten Rabatte in Euro
    """
    # Event laden
    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        return 0.0

    # Aktives Ruleset für das Event finden
    ruleset = db.query(Ruleset).filter(
        Ruleset.event_id == event_id,
        Ruleset.is_active == True,
        Ruleset.valid_from <= event.start_date,
        Ruleset.valid_until >= event.start_date
    ).first()

    if not ruleset or not ruleset.role_discounts:
        return 0.0

    # Alle aktiven Teilnehmer mit Rollen laden
    participants = db.query(Participant).filter(
        Participant.event_id == event_id,
        Participant.is_active == True,
        Participant.role_id.isnot(None)
    ).all()

    total_non_subsidy_discount = 0.0

    for participant in participants:
        if not participant.role:
            continue

        # Rolle im Regelwerk finden
        role_name_lower = participant.role.name.lower()
        role_config = None

        for key, value in ruleset.role_discounts.items():
            if key.lower() == role_name_lower:
                role_config = value
                break

        if not role_config:
            continue

        # Prüfen ob Rolle zuschussberechtigt ist (Standard: true)
        subsidy_eligible = role_config.get("subsidy_eligible", True)

        # Nur nicht-zuschussberechtigte Rollen berücksichtigen
        if subsidy_eligible:
            continue

        # Alter berechnen
        age = event.start_date.year - participant.birth_date.year
        if (event.start_date.month, event.start_date.day) < (participant.birth_date.month, participant.birth_date.day):
            age -= 1

        # Basispreis ermitteln
        base_price = PriceCalculator._get_base_price_by_age(
            age,
            ruleset.age_groups or []
        )

        # Rabatt berechnen
        discount_percent = role_config.get("discount_percent", 0)
        discount_amount = base_price * (discount_percent / 100)

        total_non_subsidy_discount += discount_amount

    return round(total_non_subsidy_discount, 2)


def calculate_expected_subsidies(db: Session, event_id: int) -> float:
    """
    Berechnet die erwarteten Zuschüsse (rollenbasiert + Familienrabatt).
    Diese Berechnung ist identisch mit der Summe im Zuschüsse-Tab.

    Berücksichtigt nur:
    - Rollen mit subsidy_eligible=true
    - Teilnehmer ohne manual_price_override
    - Kinder unter 18 für Familienrabatte

    Args:
        db: Datenbank-Session
        event_id: ID des Events

    Returns:
        Summe aller erwarteten Zuschüsse in Euro
    """
    # Event laden
    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        return 0.0

    # Aktives Ruleset für das Event finden
    ruleset = db.query(Ruleset).filter(
        Ruleset.event_id == event_id,
        Ruleset.is_active == True,
        Ruleset.valid_from <= event.start_date,
        Ruleset.valid_until >= event.start_date
    ).first()

    if not ruleset:
        return 0.0

    total_subsidies = 0.0

    # === Rollenbasierte Zuschüsse ===
    if ruleset.role_discounts:
        for role_name, role_config in ruleset.role_discounts.items():
            subsidy_eligible = role_config.get("subsidy_eligible", True)

            if not subsidy_eligible:
                continue

            # Rolle aus Datenbank laden
            role = db.query(Role).filter(
                Role.event_id == event_id,
                Role.is_active == True,
                func.lower(Role.name) == role_name.lower()
            ).first()

            if not role:
                continue

            # Teilnehmer mit dieser Rolle laden (ohne manuelle Preisanpassungen)
            participants = db.query(Participant).filter(
                Participant.event_id == event_id,
                Participant.is_active == True,
                Participant.role_id == role.id,
                Participant.manual_price_override.is_(None)
            ).all()

            discount_percent = role_config.get("discount_percent", 0)

            for participant in participants:
                # Alter berechnen
                age = event.start_date.year - participant.birth_date.year
                if (event.start_date.month, event.start_date.day) < (participant.birth_date.month, participant.birth_date.day):
                    age -= 1

                # Basispreis ermitteln
                base_price = PriceCalculator._get_base_price_by_age(
                    age,
                    ruleset.age_groups or []
                )

                # Rollenrabatt berechnen
                subsidy_amount = base_price * (discount_percent / 100)
                total_subsidies += subsidy_amount

    # === Familienrabatte (Kinderrabatt) ===
    if ruleset.family_discount and ruleset.family_discount.get("enabled", False):
        # Alle Kinder (unter 18) mit Familienrabatt laden (ohne manuelle Preisanpassungen)
        participants = db.query(Participant).filter(
            Participant.event_id == event_id,
            Participant.is_active == True,
            Participant.family_id.isnot(None),
            Participant.manual_price_override.is_(None)
        ).all()

        for participant in participants:
            # Alter berechnen
            age = event.start_date.year - participant.birth_date.year
            if (event.start_date.month, event.start_date.day) < (participant.birth_date.month, participant.birth_date.day):
                age -= 1

            # Nur Kinder unter 18
            if age >= 18:
                continue

            # Basispreis ermitteln
            base_price = PriceCalculator._get_base_price_by_age(
                age,
                ruleset.age_groups or []
            )

            # Position in Familie ermitteln
            siblings = db.query(Participant).filter(
                Participant.family_id == participant.family_id,
                Participant.is_active == True,
                Participant.event_id == event_id
            ).order_by(Participant.birth_date).all()

            # Position des Kindes bestimmen
            child_position = 1
            for idx, sibling in enumerate(siblings, start=1):
                if sibling.id == participant.id:
                    child_position = idx
                    break

            # Familienrabatt berechnen
            family_discount_percent = PriceCalculator._get_family_discount(
                age,
                child_position,
                ruleset.family_discount
            )

            subsidy_amount = base_price * (family_discount_percent / 100)

            if subsidy_amount > 0:
                total_subsidies += subsidy_amount

    return round(total_subsidies, 2)


@router.get("/", response_class=HTMLResponse)
async def cash_status(
    request: Request,
    db: Session = Depends(get_db),
    event_id: int = Depends(get_current_event_id)
):
    """Zeigt den aktuellen Kassenstand mit detaillierter Aufschlüsselung"""

    # === SOLL-Werte (Zu erwartende Werte) ===

    # Basispreise ohne Rabatte berechnen
    base_prices_sum = calculate_base_prices_sum(db, event_id)

    # Erwartete Einnahmen durch Teilnehmer (mit Rabatten/manuellen Preisen)
    # Konvertiere zu float um Decimal/float Typ-Konflikte zu vermeiden
    participants = db.query(Participant).filter(
        Participant.event_id == event_id,
        Participant.is_active == True
    ).all()
    expected_income_participants = float(sum((p.final_price for p in participants), 0))

    # Erwartete Zuschüsse berechnen (rollenbasiert + Familienrabatt)
    # WICHTIG: Diese Berechnung muss identisch sein mit der Summe im Zuschüsse-Tab!
    expected_subsidies = calculate_expected_subsidies(db, event_id)

    # Rabatte für nicht-zuschussberechtigte Rollen berechnen (werden auf Gruppe umgelegt)
    non_subsidy_discounts = calculate_non_subsidy_discount_sum(db, event_id)

    # Gesamteinnahmen (Soll) = Basispreise MINUS nicht-zuschussberechtigte Rabatte
    # (Nicht-zuschussberechtigte Rabatte sind Umlagen auf die Gruppe, keine erwarteten Einnahmen)
    expected_total_income = base_prices_sum - non_subsidy_discounts

    # Sonstige Einnahmen (SOLL) = Erwartete Zuschüsse
    # Diese Summe muss identisch sein mit der Summe aller Zuschüsse im Zuschüsse-Tab
    other_income = expected_subsidies

    # Alle Ausgaben (gesamt)
    total_expenses = float(db.query(func.sum(Expense.amount)).filter(
        Expense.event_id == event_id
    ).scalar() or 0)

    # Erwarteter Saldo (basierend auf erwarteten Gesamteinnahmen)
    expected_balance = expected_total_income - total_expenses

    # === IST-Werte (Getätigte Zahlungen) ===

    # Tatsächliche Einnahmen durch Teilnehmer-Zahlungen
    actual_income_participants = float(db.query(func.sum(Payment.amount)).filter(
        Payment.event_id == event_id
    ).scalar() or 0)

    # Sonstige Einnahmen (z.B. erhaltene Zuschüsse, Spenden)
    # Aus Income-Tabelle abfragen
    actual_other_income = float(db.query(func.sum(Income.amount)).filter(
        Income.event_id == event_id
    ).scalar() or 0)

    # Beglichene Ausgaben
    settled_expenses = float(db.query(func.sum(Expense.amount)).filter(
        Expense.event_id == event_id,
        Expense.is_settled == True
    ).scalar() or 0)

    # Aktueller Saldo
    actual_balance = actual_income_participants + actual_other_income - settled_expenses

    # === DIFFERENZEN ===

    # Ausstehende Einnahmen (Teilnehmer)
    outstanding_income_participants = expected_income_participants - actual_income_participants

    # Ausstehende sonstige Einnahmen (erwartete Zuschüsse minus erhaltene)
    outstanding_other_income = other_income - actual_other_income

    # Noch zu begleichende Ausgaben
    open_expenses = float(db.query(func.sum(Expense.amount)).filter(
        Expense.event_id == event_id,
        Expense.is_settled == False
    ).scalar() or 0)

    # Differenz Saldo
    balance_difference = expected_balance - actual_balance

    # Status-Badge berechnen (basierend auf aktuellem Saldo)
    if actual_balance < 0:
        status = {"text": "Kritisch", "color": "red"}
    elif actual_balance < 500:
        status = {"text": "Knapp", "color": "yellow"}
    else:
        status = {"text": "Gesund", "color": "green"}

    # Ausgaben nach Kategorie
    expenses_by_category = db.query(
        Expense.category,
        func.sum(Expense.amount).label("total"),
        func.count(Expense.id).label("count")
    ).filter(
        Expense.event_id == event_id
    ).group_by(Expense.category).all()

    # Kategorien aufbereiten
    categories = []
    for cat, total, count in expenses_by_category:
        categories.append({
            "name": cat if cat else "Sonstige",
            "total": total,
            "count": count
        })

    # Nach Betrag sortieren
    categories.sort(key=lambda x: x["total"], reverse=True)

    return templates.TemplateResponse(
        "cash_status/overview.html",
        {
            "request": request,
            "title": "Kassenstand",
            # SOLL-Werte
            "expected_income_participants": expected_income_participants,
            "expected_other_income": other_income,
            "expected_total_income": expected_total_income,
            "expected_expenses": total_expenses,
            "expected_balance": expected_balance,
            # IST-Werte
            "actual_income_participants": actual_income_participants,
            "actual_other_income": actual_other_income,
            "actual_expenses": settled_expenses,
            "actual_balance": actual_balance,
            # DIFFERENZEN
            "outstanding_income_participants": outstanding_income_participants,
            "outstanding_other_income": outstanding_other_income,
            "outstanding_expenses": open_expenses,
            "balance_difference": balance_difference,
            # Status und Kategorien
            "status": status,
            "categories": categories
        }
    )


@router.get("/history", response_class=HTMLResponse)
async def transaction_history(
    request: Request,
    db: Session = Depends(get_db),
    event_id: int = Depends(get_current_event_id),
    date_from: Optional[str] = Query(None, description="Startdatum (YYYY-MM-DD)"),
    date_to: Optional[str] = Query(None, description="Enddatum (YYYY-MM-DD)"),
    transaction_type: Optional[str] = Query(None, description="Transaktionstyp (payment/income/expense)"),
    min_amount: Optional[float] = Query(None, description="Minimalbetrag"),
    max_amount: Optional[float] = Query(None, description="Maximalbetrag"),
    search: Optional[str] = Query(None, description="Suchbegriff"),
):
    """
    Zeigt die vollständige Transaktionshistorie mit allen Ein- und Ausgängen
    """
    logger.info(f"Loading transaction history for event {event_id} with filters")

    # === Query für Zahlungseingänge (Payments) ===
    payments_query = db.query(
        Payment.id.label('id'),
        literal('payment').label('type'),
        Payment.amount.label('amount'),
        Payment.payment_date.label('transaction_date'),
        Payment.payment_method.label('method'),
        Payment.reference.label('reference'),
        Payment.notes.label('description'),
        Participant.first_name.label('participant_first_name'),
        Participant.last_name.label('participant_last_name'),
        Family.name.label('family_name'),
        Payment.created_at.label('created_at')
    ).outerjoin(
        Participant, Payment.participant_id == Participant.id
    ).outerjoin(
        Family, Payment.family_id == Family.id
    ).filter(
        Payment.event_id == event_id
    )

    # === Query für sonstige Einnahmen (Incomes) ===
    incomes_query = db.query(
        Income.id.label('id'),
        literal('income').label('type'),
        Income.amount.label('amount'),
        Income.date.label('transaction_date'),
        literal(None).label('method'),
        Income.name.label('reference'),
        Income.description.label('description'),
        literal(None).label('participant_first_name'),
        literal(None).label('participant_last_name'),
        literal(None).label('family_name'),
        literal(None).label('created_at')
    ).filter(
        Income.event_id == event_id
    )

    # === Query für Ausgaben (Expenses) ===
    expenses_query = db.query(
        Expense.id.label('id'),
        literal('expense').label('type'),
        (Expense.amount * -1).label('amount'),  # Negativ für Ausgaben
        Expense.expense_date.label('transaction_date'),
        Expense.category.label('method'),
        Expense.title.label('reference'),
        Expense.description.label('description'),
        literal(None).label('participant_first_name'),
        literal(None).label('participant_last_name'),
        literal(None).label('family_name'),
        Expense.created_at.label('created_at')
    ).filter(
        Expense.event_id == event_id
    )

    # === Filter anwenden ===

    # Datumsfilter
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, "%Y-%m-%d").date()
            payments_query = payments_query.filter(Payment.payment_date >= date_from_obj)
            incomes_query = incomes_query.filter(Income.date >= date_from_obj)
            expenses_query = expenses_query.filter(Expense.expense_date >= date_from_obj)
        except ValueError:
            logger.warning(f"Invalid date_from format: {date_from}")

    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, "%Y-%m-%d").date()
            payments_query = payments_query.filter(Payment.payment_date <= date_to_obj)
            incomes_query = incomes_query.filter(Income.date <= date_to_obj)
            expenses_query = expenses_query.filter(Expense.expense_date <= date_to_obj)
        except ValueError:
            logger.warning(f"Invalid date_to format: {date_to}")

    # Betragsfilter
    if min_amount is not None:
        payments_query = payments_query.filter(Payment.amount >= min_amount)
        incomes_query = incomes_query.filter(Income.amount >= min_amount)
        expenses_query = expenses_query.filter(Expense.amount >= min_amount)

    if max_amount is not None:
        payments_query = payments_query.filter(Payment.amount <= max_amount)
        incomes_query = incomes_query.filter(Income.amount <= max_amount)
        expenses_query = expenses_query.filter(Expense.amount <= max_amount)

    # Typ-Filter
    queries_to_union = []
    if not transaction_type or transaction_type == 'payment':
        queries_to_union.append(payments_query)
    if not transaction_type or transaction_type == 'income':
        queries_to_union.append(incomes_query)
    if not transaction_type or transaction_type == 'expense':
        queries_to_union.append(expenses_query)

    # Union aller Queries
    if queries_to_union:
        combined_query = union_all(*queries_to_union)
        transactions = db.execute(combined_query).fetchall()
    else:
        transactions = []

    # In Liste von Dictionaries umwandeln für einfachere Verarbeitung
    transactions_list = []
    for t in transactions:
        transaction_dict = {
            'id': t.id,
            'type': t.type,
            'amount': float(t.amount),  # Konvertiere Decimal zu float
            'transaction_date': t.transaction_date,
            'method': t.method,
            'reference': t.reference,
            'description': t.description,
            'participant_name': f"{t.participant_first_name} {t.participant_last_name}" if t.participant_first_name else None,
            'family_name': t.family_name,
            'created_at': t.created_at
        }
        transactions_list.append(transaction_dict)

    # Suchfilter (auf Liste anwenden)
    if search:
        search_lower = search.lower()
        transactions_list = [
            t for t in transactions_list
            if (t['reference'] and search_lower in t['reference'].lower()) or
               (t['description'] and search_lower in t['description'].lower()) or
               (t['participant_name'] and search_lower in t['participant_name'].lower()) or
               (t['family_name'] and search_lower in t['family_name'].lower())
        ]

    # Nach Datum sortieren (neueste zuerst)
    transactions_list.sort(key=lambda x: x['transaction_date'], reverse=True)

    # Summen berechnen
    total_income = sum(t['amount'] for t in transactions_list if t['amount'] > 0)
    total_expenses = sum(abs(t['amount']) for t in transactions_list if t['amount'] < 0)
    net_total = total_income - total_expenses

    # === Erweiterte Kategorisierung ===
    # Gruppierung nach Typ und Kategorie/Methode
    categories_summary = {}

    for t in transactions_list:
        if t['amount'] > 0:
            # Einnahmen nach Typ kategorisieren
            category_key = t['type']
            if category_key not in categories_summary:
                categories_summary[category_key] = {
                    'name': category_key,
                    'type': 'income',
                    'total': 0.0,
                    'count': 0,
                    'transactions': []
                }
            categories_summary[category_key]['total'] += t['amount']
            categories_summary[category_key]['count'] += 1
            categories_summary[category_key]['transactions'].append(t)
        else:
            # Ausgaben nach Kategorie/Methode
            category_key = t['method'] if t['method'] else 'Sonstiges'
            if category_key not in categories_summary:
                categories_summary[category_key] = {
                    'name': category_key,
                    'type': 'expense',
                    'total': 0.0,
                    'count': 0,
                    'transactions': []
                }
            categories_summary[category_key]['total'] += abs(t['amount'])
            categories_summary[category_key]['count'] += 1
            categories_summary[category_key]['transactions'].append(t)

    # Sortieren nach Betrag (höchste zuerst)
    categories_list = sorted(categories_summary.values(), key=lambda x: x['total'], reverse=True)

    # Separate Listen für Einnahmen und Ausgaben
    income_categories = [c for c in categories_list if c['type'] == 'income']
    expense_categories = [c for c in categories_list if c['type'] == 'expense']

    # Laufenden Saldo berechnen (chronologisch)
    transactions_with_balance = []
    running_balance = 0.0
    for t in reversed(transactions_list):  # Chronologisch
        running_balance += t['amount']
        t['running_balance'] = running_balance
        transactions_with_balance.append(t)

    transactions_with_balance.reverse()  # Wieder zurück zu neueste zuerst

    logger.info(f"Loaded {len(transactions_with_balance)} transactions in {len(categories_list)} categories")

    return templates.TemplateResponse(
        "cash_status/overview.html",
        {
            "request": request,
            "title": "Kassenstand - Historie",
            "show_history": True,
            "transactions": transactions_with_balance,
            "total_income": total_income,
            "total_expenses": total_expenses,
            "net_total": net_total,
            # Kategorisierung
            "income_categories": income_categories,
            "expense_categories": expense_categories,
            # Filter-Werte
            "filter_date_from": date_from,
            "filter_date_to": date_to,
            "filter_type": transaction_type,
            "filter_min_amount": min_amount,
            "filter_max_amount": max_amount,
            "filter_search": search,
        }
    )


@router.get("/history/export/excel")
async def export_history_excel(
    db: Session = Depends(get_db),
    event_id: int = Depends(get_current_event_id),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    transaction_type: Optional[str] = Query(None),
    min_amount: Optional[float] = Query(None),
    max_amount: Optional[float] = Query(None),
    search: Optional[str] = Query(None),
):
    """Exportiert die Transaktionshistorie als Excel-Datei"""
    logger.info(f"Exporting transaction history to Excel for event {event_id}")

    # Transaktionen mit denselben Filtern wie in der Historie laden
    # (Wiederverwendung der Logik aus transaction_history)

    # === Queries aufbauen ===
    payments_query = db.query(
        Payment.id.label('id'),
        literal('Zahlungseingang').label('type'),
        Payment.amount.label('amount'),
        Payment.payment_date.label('transaction_date'),
        Payment.payment_method.label('method'),
        Payment.reference.label('reference'),
        Payment.notes.label('description'),
        Participant.first_name.label('participant_first_name'),
        Participant.last_name.label('participant_last_name'),
        Family.name.label('family_name')
    ).outerjoin(
        Participant, Payment.participant_id == Participant.id
    ).outerjoin(
        Family, Payment.family_id == Family.id
    ).filter(
        Payment.event_id == event_id
    )

    incomes_query = db.query(
        Income.id.label('id'),
        literal('Sonstige Einnahme').label('type'),
        Income.amount.label('amount'),
        Income.date.label('transaction_date'),
        literal(None).label('method'),
        Income.name.label('reference'),
        Income.description.label('description'),
        literal(None).label('participant_first_name'),
        literal(None).label('participant_last_name'),
        literal(None).label('family_name')
    ).filter(
        Income.event_id == event_id
    )

    expenses_query = db.query(
        Expense.id.label('id'),
        literal('Ausgabe').label('type'),
        (Expense.amount * -1).label('amount'),
        Expense.expense_date.label('transaction_date'),
        Expense.category.label('method'),
        Expense.title.label('reference'),
        Expense.description.label('description'),
        literal(None).label('participant_first_name'),
        literal(None).label('participant_last_name'),
        literal(None).label('family_name')
    ).filter(
        Expense.event_id == event_id
    )

    # Filter anwenden
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, "%Y-%m-%d").date()
            payments_query = payments_query.filter(Payment.payment_date >= date_from_obj)
            incomes_query = incomes_query.filter(Income.date >= date_from_obj)
            expenses_query = expenses_query.filter(Expense.expense_date >= date_from_obj)
        except ValueError:
            pass

    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, "%Y-%m-%d").date()
            payments_query = payments_query.filter(Payment.payment_date <= date_to_obj)
            incomes_query = incomes_query.filter(Income.date <= date_to_obj)
            expenses_query = expenses_query.filter(Expense.expense_date <= date_to_obj)
        except ValueError:
            pass

    if min_amount is not None:
        payments_query = payments_query.filter(Payment.amount >= min_amount)
        incomes_query = incomes_query.filter(Income.amount >= min_amount)
        expenses_query = expenses_query.filter(Expense.amount >= min_amount)

    if max_amount is not None:
        payments_query = payments_query.filter(Payment.amount <= max_amount)
        incomes_query = incomes_query.filter(Income.amount <= max_amount)
        expenses_query = expenses_query.filter(Expense.amount <= max_amount)

    # Union
    queries_to_union = []
    if not transaction_type or transaction_type == 'payment':
        queries_to_union.append(payments_query)
    if not transaction_type or transaction_type == 'income':
        queries_to_union.append(incomes_query)
    if not transaction_type or transaction_type == 'expense':
        queries_to_union.append(expenses_query)

    if queries_to_union:
        combined_query = union_all(*queries_to_union)
        transactions = db.execute(combined_query).fetchall()
    else:
        transactions = []

    # In Liste umwandeln
    transactions_list = []
    for t in transactions:
        transaction_dict = {
            'type': t.type,
            'transaction_date': t.transaction_date,
            'amount': t.amount,
            'method': t.method,
            'reference': t.reference,
            'description': t.description,
            'participant_name': f"{t.participant_first_name} {t.participant_last_name}" if t.participant_first_name else None,
            'family_name': t.family_name,
        }
        transactions_list.append(transaction_dict)

    # Suchfilter
    if search:
        search_lower = search.lower()
        transactions_list = [
            t for t in transactions_list
            if (t['reference'] and search_lower in t['reference'].lower()) or
               (t['description'] and search_lower in t['description'].lower()) or
               (t['participant_name'] and search_lower in t['participant_name'].lower()) or
               (t['family_name'] and search_lower in t['family_name'].lower())
        ]

    # Chronologisch sortieren (älteste zuerst für Excel)
    transactions_list.sort(key=lambda x: x['transaction_date'])

    # Excel erstellen mit ExcelService
    wb, ws = ExcelService.create_workbook("Transaktionshistorie")

    # Header
    headers = [
        "Datum", "Typ", "Betrag (€)", "Einnahme (€)", "Ausgabe (€)",
        "Saldo (€)", "Kategorie/Methode", "Referenz", "Beschreibung",
        "Teilnehmer", "Familie"
    ]

    # Spaltenbreiten
    column_widths = {1: 12, 2: 18, 3: 12, 4: 12, 5: 12, 6: 12, 7: 15, 8: 25, 9: 30, 10: 20, 11: 20}

    # Header-Formatierung mit ExcelService
    ExcelService.apply_header_row(ws, headers, column_widths)

    # Daten schreiben
    running_balance = 0.0
    row_num = 2

    for transaction in transactions_list:
        running_balance += transaction['amount']

        ws.cell(row=row_num, column=1, value=transaction['transaction_date'].strftime('%d.%m.%Y'))
        ws.cell(row=row_num, column=2, value=transaction['type'])
        ws.cell(row=row_num, column=3, value=transaction['amount'])
        ws.cell(row=row_num, column=4, value=transaction['amount'] if transaction['amount'] > 0 else 0)
        ws.cell(row=row_num, column=5, value=abs(transaction['amount']) if transaction['amount'] < 0 else 0)
        ws.cell(row=row_num, column=6, value=running_balance)
        ws.cell(row=row_num, column=7, value=transaction['method'] or '')
        ws.cell(row=row_num, column=8, value=transaction['reference'] or '')
        ws.cell(row=row_num, column=9, value=transaction['description'] or '')
        ws.cell(row=row_num, column=10, value=transaction['participant_name'] or '')
        ws.cell(row=row_num, column=11, value=transaction['family_name'] or '')

        # Farbe für Einnahmen/Ausgaben mit ExcelService
        ExcelService.apply_color_by_value(ws, row_num, 3, transaction['amount'])

        row_num += 1

    # Summenzeile mit ExcelService
    row_num += 1
    total_income = sum(t['amount'] for t in transactions_list if t['amount'] > 0)
    total_expenses = sum(abs(t['amount']) for t in transactions_list if t['amount'] < 0)

    summary_values = {
        4: total_income,
        5: total_expenses,
        6: running_balance
    }
    ExcelService.apply_summary_row(ws, row_num, summary_values)

    # Datei als BytesIO zurückgeben
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Transaktionshistorie_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    logger.info(f"Excel export completed: {len(transactions_list)} transactions")

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/history/export/csv")
async def export_history_csv(
    db: Session = Depends(get_db),
    event_id: int = Depends(get_current_event_id),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    transaction_type: Optional[str] = Query(None),
    min_amount: Optional[float] = Query(None),
    max_amount: Optional[float] = Query(None),
    search: Optional[str] = Query(None),
):
    """Exportiert die Transaktionshistorie als CSV-Datei"""
    logger.info(f"Exporting transaction history to CSV for event {event_id}")

    # Queries aufbauen (wie bei Excel)
    payments_query = db.query(
        literal('Zahlungseingang').label('type'),
        Payment.payment_date.label('transaction_date'),
        Payment.amount.label('amount'),
        Payment.payment_method.label('method'),
        Payment.reference.label('reference'),
        Payment.notes.label('description'),
        Participant.first_name.label('participant_first_name'),
        Participant.last_name.label('participant_last_name'),
        Family.name.label('family_name')
    ).outerjoin(
        Participant, Payment.participant_id == Participant.id
    ).outerjoin(
        Family, Payment.family_id == Family.id
    ).filter(
        Payment.event_id == event_id
    )

    incomes_query = db.query(
        literal('Sonstige Einnahme').label('type'),
        Income.date.label('transaction_date'),
        Income.amount.label('amount'),
        literal(None).label('method'),
        Income.name.label('reference'),
        Income.description.label('description'),
        literal(None).label('participant_first_name'),
        literal(None).label('participant_last_name'),
        literal(None).label('family_name')
    ).filter(
        Income.event_id == event_id
    )

    expenses_query = db.query(
        literal('Ausgabe').label('type'),
        Expense.expense_date.label('transaction_date'),
        (Expense.amount * -1).label('amount'),
        Expense.category.label('method'),
        Expense.title.label('reference'),
        Expense.description.label('description'),
        literal(None).label('participant_first_name'),
        literal(None).label('participant_last_name'),
        literal(None).label('family_name')
    ).filter(
        Expense.event_id == event_id
    )

    # Filter anwenden
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, "%Y-%m-%d").date()
            payments_query = payments_query.filter(Payment.payment_date >= date_from_obj)
            incomes_query = incomes_query.filter(Income.date >= date_from_obj)
            expenses_query = expenses_query.filter(Expense.expense_date >= date_from_obj)
        except ValueError:
            pass

    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, "%Y-%m-%d").date()
            payments_query = payments_query.filter(Payment.payment_date <= date_to_obj)
            incomes_query = incomes_query.filter(Income.date <= date_to_obj)
            expenses_query = expenses_query.filter(Expense.expense_date <= date_to_obj)
        except ValueError:
            pass

    if min_amount is not None:
        payments_query = payments_query.filter(Payment.amount >= min_amount)
        incomes_query = incomes_query.filter(Income.amount >= min_amount)
        expenses_query = expenses_query.filter(Expense.amount >= min_amount)

    if max_amount is not None:
        payments_query = payments_query.filter(Payment.amount <= max_amount)
        incomes_query = incomes_query.filter(Income.amount <= max_amount)
        expenses_query = expenses_query.filter(Expense.amount <= max_amount)

    # Union
    queries_to_union = []
    if not transaction_type or transaction_type == 'payment':
        queries_to_union.append(payments_query)
    if not transaction_type or transaction_type == 'income':
        queries_to_union.append(incomes_query)
    if not transaction_type or transaction_type == 'expense':
        queries_to_union.append(expenses_query)

    if queries_to_union:
        combined_query = union_all(*queries_to_union)
        transactions = db.execute(combined_query).fetchall()
    else:
        transactions = []

    # In Liste umwandeln
    transactions_list = []
    for t in transactions:
        transaction_dict = {
            'type': t.type,
            'transaction_date': t.transaction_date,
            'amount': t.amount,
            'method': t.method or '',
            'reference': t.reference or '',
            'description': t.description or '',
            'participant_name': f"{t.participant_first_name} {t.participant_last_name}" if t.participant_first_name else '',
            'family_name': t.family_name or '',
        }
        transactions_list.append(transaction_dict)

    # Suchfilter
    if search:
        search_lower = search.lower()
        transactions_list = [
            t for t in transactions_list
            if (search_lower in t['reference'].lower()) or
               (search_lower in t['description'].lower()) or
               (search_lower in t['participant_name'].lower()) or
               (search_lower in t['family_name'].lower())
        ]

    # Chronologisch sortieren
    transactions_list.sort(key=lambda x: x['transaction_date'])

    # CSV erstellen
    output = StringIO()
    writer = csv.writer(output, delimiter=';', quoting=csv.QUOTE_MINIMAL)

    # Header
    writer.writerow([
        "Datum", "Typ", "Betrag", "Einnahme", "Ausgabe", "Saldo",
        "Kategorie/Methode", "Referenz", "Beschreibung", "Teilnehmer", "Familie"
    ])

    # Daten
    running_balance = 0.0
    for transaction in transactions_list:
        running_balance += transaction['amount']

        writer.writerow([
            transaction['transaction_date'].strftime('%d.%m.%Y'),
            transaction['type'],
            f"{transaction['amount']:.2f}",
            f"{transaction['amount']:.2f}" if transaction['amount'] > 0 else "0.00",
            f"{abs(transaction['amount']):.2f}" if transaction['amount'] < 0 else "0.00",
            f"{running_balance:.2f}",
            transaction['method'],
            transaction['reference'],
            transaction['description'],
            transaction['participant_name'],
            transaction['family_name']
        ])

    filename = f"Transaktionshistorie_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

    logger.info(f"CSV export completed: {len(transactions_list)} transactions")

    return Response(
        content=output.getvalue().encode('utf-8-sig'),  # BOM für Excel
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/history/export/pdf")
async def export_history_pdf(
    db: Session = Depends(get_db),
    event_id: int = Depends(get_current_event_id),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    transaction_type: Optional[str] = Query(None),
    min_amount: Optional[float] = Query(None),
    max_amount: Optional[float] = Query(None),
    search: Optional[str] = Query(None),
):
    """Exportiert die Transaktionshistorie als PDF-Datei für Steuerberater/Nachweispflicht"""
    logger.info(f"Exporting transaction history to PDF for event {event_id}")

    # Event-Daten laden
    event = db.query(Event).filter(Event.id == event_id).first()
    event_name = event.name if event else f"Event {event_id}"

    # === Queries aufbauen (gleiche Logik wie Excel/CSV) ===
    payments_query = db.query(
        Payment.id.label('id'),
        literal('Zahlungseingang').label('type'),
        Payment.amount.label('amount'),
        Payment.payment_date.label('transaction_date'),
        Payment.payment_method.label('method'),
        Payment.reference.label('reference'),
        Payment.notes.label('description'),
        Participant.first_name.label('participant_first_name'),
        Participant.last_name.label('participant_last_name'),
        Family.name.label('family_name'),
        literal(None).label('receipt_available')  # Zahlungseingänge haben keine Belege
    ).outerjoin(
        Participant, Payment.participant_id == Participant.id
    ).outerjoin(
        Family, Payment.family_id == Family.id
    ).filter(
        Payment.event_id == event_id
    )

    incomes_query = db.query(
        Income.id.label('id'),
        literal('Sonstige Einnahme').label('type'),
        Income.amount.label('amount'),
        Income.date.label('transaction_date'),
        literal(None).label('method'),
        Income.name.label('reference'),
        Income.description.label('description'),
        literal(None).label('participant_first_name'),
        literal(None).label('participant_last_name'),
        literal(None).label('family_name'),
        Income.receipt_file_path.label('receipt_available')
    ).filter(
        Income.event_id == event_id
    )

    expenses_query = db.query(
        Expense.id.label('id'),
        literal('Ausgabe').label('type'),
        (Expense.amount * -1).label('amount'),
        Expense.expense_date.label('transaction_date'),
        Expense.category.label('method'),
        Expense.title.label('reference'),
        Expense.description.label('description'),
        literal(None).label('participant_first_name'),
        literal(None).label('participant_last_name'),
        literal(None).label('family_name'),
        Expense.receipt_file_path.label('receipt_available')
    ).filter(
        Expense.event_id == event_id
    )

    # Filter anwenden
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, "%Y-%m-%d").date()
            payments_query = payments_query.filter(Payment.payment_date >= date_from_obj)
            incomes_query = incomes_query.filter(Income.date >= date_from_obj)
            expenses_query = expenses_query.filter(Expense.expense_date >= date_from_obj)
        except ValueError:
            pass

    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, "%Y-%m-%d").date()
            payments_query = payments_query.filter(Payment.payment_date <= date_to_obj)
            incomes_query = incomes_query.filter(Income.date <= date_to_obj)
            expenses_query = expenses_query.filter(Expense.expense_date <= date_to_obj)
        except ValueError:
            pass

    if min_amount is not None:
        payments_query = payments_query.filter(Payment.amount >= min_amount)
        incomes_query = incomes_query.filter(Income.amount >= min_amount)
        expenses_query = expenses_query.filter(Expense.amount >= min_amount)

    if max_amount is not None:
        payments_query = payments_query.filter(Payment.amount <= max_amount)
        incomes_query = incomes_query.filter(Income.amount <= max_amount)
        expenses_query = expenses_query.filter(Expense.amount <= max_amount)

    # Union
    queries_to_union = []
    if not transaction_type or transaction_type == 'payment':
        queries_to_union.append(payments_query)
    if not transaction_type or transaction_type == 'income':
        queries_to_union.append(incomes_query)
    if not transaction_type or transaction_type == 'expense':
        queries_to_union.append(expenses_query)

    if queries_to_union:
        combined_query = union_all(*queries_to_union)
        transactions = db.execute(combined_query).fetchall()
    else:
        transactions = []

    # In Liste umwandeln
    transactions_list = []
    for t in transactions:
        transaction_dict = {
            'type': t.type,
            'transaction_date': t.transaction_date,
            'amount': t.amount,
            'method': t.method,
            'reference': t.reference,
            'description': t.description,
            'participant_name': f"{t.participant_first_name} {t.participant_last_name}" if t.participant_first_name else None,
            'family_name': t.family_name,
            'receipt_available': t.receipt_available,
        }
        transactions_list.append(transaction_dict)

    # Suchfilter
    if search:
        search_lower = search.lower()
        transactions_list = [
            t for t in transactions_list
            if (t['reference'] and search_lower in t['reference'].lower()) or
               (t['description'] and search_lower in t['description'].lower()) or
               (t['participant_name'] and search_lower in t['participant_name'].lower()) or
               (t['family_name'] and search_lower in t['family_name'].lower())
        ]

    # Chronologisch sortieren (älteste zuerst)
    transactions_list.sort(key=lambda x: x['transaction_date'])

    # Summen berechnen
    total_income = sum(t['amount'] for t in transactions_list if t['amount'] > 0)
    total_expenses = sum(abs(t['amount']) for t in transactions_list if t['amount'] < 0)
    net_total = total_income - total_expenses

    # === PDF erstellen ===
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        topMargin=1.5*cm,
        bottomMargin=2*cm,
        leftMargin=1.5*cm,
        rightMargin=1.5*cm
    )
    story = []

    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=20,
        alignment=1  # Center
    )
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=12,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=10,
    )
    normal_style = styles['Normal']
    small_style = ParagraphStyle(
        'Small',
        parent=styles['Normal'],
        fontSize=8,
    )

    # === Header ===
    story.append(Paragraph(f"Transaktionshistorie", title_style))
    story.append(Paragraph(f"Event: {event_name}", heading_style))

    # Zeitraum und Filter-Informationen
    export_date = datetime.now().strftime("%d.%m.%Y %H:%M")
    info_text = f"Exportiert am: {export_date}<br/>"

    if date_from or date_to:
        period = "Zeitraum: "
        if date_from:
            period += f"ab {datetime.strptime(date_from, '%Y-%m-%d').strftime('%d.%m.%Y')}"
        if date_to:
            if date_from:
                period += f" bis {datetime.strptime(date_to, '%Y-%m-%d').strftime('%d.%m.%Y')}"
            else:
                period += f"bis {datetime.strptime(date_to, '%Y-%m-%d').strftime('%d.%m.%Y')}"
        info_text += period + "<br/>"

    if transaction_type:
        type_names = {
            'payment': 'Zahlungseingänge',
            'income': 'Sonstige Einnahmen',
            'expense': 'Ausgaben'
        }
        info_text += f"Filter: {type_names.get(transaction_type, transaction_type)}<br/>"

    if min_amount is not None or max_amount is not None:
        amount_filter = "Betrag: "
        if min_amount is not None:
            amount_filter += f"ab {min_amount:.2f} €"
        if max_amount is not None:
            if min_amount is not None:
                amount_filter += f" bis {max_amount:.2f} €"
            else:
                amount_filter += f"bis {max_amount:.2f} €"
        info_text += amount_filter + "<br/>"

    if search:
        info_text += f"Suchbegriff: '{search}'<br/>"

    story.append(Paragraph(info_text, small_style))
    story.append(Spacer(1, 0.5*cm))

    # === Zusammenfassung ===
    summary_data = [
        ["Gesamt-Einnahmen:", f"{total_income:.2f} €"],
        ["Gesamt-Ausgaben:", f"{total_expenses:.2f} €"],
        ["Netto-Saldo:", f"{net_total:.2f} €"],
        ["Anzahl Transaktionen:", str(len(transactions_list))]
    ]

    summary_table = Table(summary_data, colWidths=[10*cm, 5*cm])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#E8F4F8')),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#1e40af')),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TEXTCOLOR', (1, 2), (1, 2), colors.HexColor('#1e40af') if net_total >= 0 else colors.red),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 0.8*cm))

    # === Transaktions-Tabelle ===
    story.append(Paragraph("Transaktionen", heading_style))

    # Tabellen-Header
    table_data = [
        ["Datum", "Typ", "Betrag", "Saldo", "Beleg", "Referenz", "Beschreibung"]
    ]

    # Daten schreiben
    running_balance = 0.0
    for transaction in transactions_list:
        running_balance += transaction['amount']

        # Beleg-Status
        receipt_status = "Ja" if transaction['receipt_available'] else "-"

        # Betrag mit Vorzeichen und Farbe
        amount_str = f"{transaction['amount']:.2f} €"

        # Beschreibung zusammenstellen
        desc_parts = []
        if transaction['reference']:
            desc_parts.append(transaction['reference'])
        if transaction['participant_name']:
            desc_parts.append(f"({transaction['participant_name']})")
        if transaction['family_name']:
            desc_parts.append(f"Familie: {transaction['family_name']}")
        if transaction['description']:
            desc_parts.append(transaction['description'])
        description_text = " - ".join(desc_parts) if desc_parts else ""

        # Beschreibung auf max. 40 Zeichen kürzen für bessere Lesbarkeit
        if len(description_text) > 40:
            description_text = description_text[:37] + "..."

        table_data.append([
            transaction['transaction_date'].strftime('%d.%m.%Y'),
            transaction['type'],
            amount_str,
            f"{running_balance:.2f} €",
            receipt_status,
            transaction['reference'] or "",
            description_text
        ])

    # Tabelle erstellen (mit automatischer Seitenumbruch-Funktion von ReportLab)
    trans_table = Table(table_data, colWidths=[2*cm, 3.2*cm, 2.2*cm, 2.2*cm, 1.5*cm, 3.5*cm, 4*cm])

    # Tabellen-Style
    table_style = [
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),

        # Daten
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ALIGN', (0, 1), (1, -1), 'LEFT'),  # Datum, Typ
        ('ALIGN', (2, 1), (4, -1), 'RIGHT'),  # Betrag, Saldo, Beleg
        ('ALIGN', (5, 1), (-1, -1), 'LEFT'),  # Referenz, Beschreibung

        # Gitter
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),

        # Alternierende Zeilen-Farben
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
    ]

    # Farben für positive/negative Beträge
    for idx, transaction in enumerate(transactions_list, start=1):
        if transaction['amount'] > 0:
            table_style.append(('TEXTCOLOR', (2, idx), (2, idx), colors.HexColor('#00B050')))
        else:
            table_style.append(('TEXTCOLOR', (2, idx), (2, idx), colors.HexColor('#C00000')))

    trans_table.setStyle(TableStyle(table_style))
    story.append(trans_table)

    # === Footer-Funktion für Seitenzahlen ===
    def add_page_number(canvas, doc):
        """Fügt Seitenzahlen und Footer hinzu"""
        page_num = canvas.getPageNumber()
        text = f"Seite {page_num} - Generiert am {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}"
        canvas.setFont('Helvetica', 8)
        canvas.setFillColor(colors.grey)
        canvas.drawRightString(A4[0] - 1.5*cm, 1*cm, text)

    # PDF generieren
    doc.build(story, onFirstPage=add_page_number, onLaterPages=add_page_number)
    buffer.seek(0)

    filename = f"Transaktionshistorie_{event_name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"

    logger.info(f"PDF export completed: {len(transactions_list)} transactions")

    return Response(
        content=buffer.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/subsidies", response_class=HTMLResponse)
async def subsidies_overview(
    request: Request,
    db: Session = Depends(get_db),
    event_id: int = Depends(get_current_event_id)
):
    """
    Zeigt die Zuschuss-Übersicht mit rollenbasierten Zuschüssen und Kinderrabatten
    """
    logger.info(f"Loading subsidies overview for event {event_id}")

    # Event laden
    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        logger.warning(f"Event {event_id} not found")
        return templates.TemplateResponse(
            "cash_status/overview.html",
            {
                "request": request,
                "title": "Zuschüsse",
                "show_subsidies": True,
                "role_subsidies": [],
                "family_subsidies": None
            }
        )

    # Aktives Ruleset für das Event finden
    ruleset = db.query(Ruleset).filter(
        Ruleset.event_id == event_id,
        Ruleset.is_active == True,
        Ruleset.valid_from <= event.start_date,
        Ruleset.valid_until >= event.start_date
    ).first()

    if not ruleset:
        logger.warning(f"No active ruleset found for event {event_id}")
        return templates.TemplateResponse(
            "cash_status/overview.html",
            {
                "request": request,
                "title": "Zuschüsse",
                "show_subsidies": True,
                "role_subsidies": [],
                "family_subsidies": None
            }
        )

    # === Rollenbasierte Zuschüsse ===
    role_subsidies = []

    # Alle Rollen mit subsidy_eligible = true finden
    if ruleset.role_discounts:
        for role_name, role_config in ruleset.role_discounts.items():
            subsidy_eligible = role_config.get("subsidy_eligible", True)

            if not subsidy_eligible:
                continue

            # Rolle aus Datenbank laden
            role = db.query(Role).filter(
                Role.event_id == event_id,
                Role.is_active == True,
                func.lower(Role.name) == role_name.lower()
            ).first()

            if not role:
                continue

            # Teilnehmer mit dieser Rolle laden (ohne manuelle Preisanpassungen)
            participants = db.query(Participant).filter(
                Participant.event_id == event_id,
                Participant.is_active == True,
                Participant.role_id == role.id,
                Participant.manual_price_override.is_(None)
            ).all()

            if not participants:
                continue

            # Für jeden Teilnehmer: Basispreis und Zuschuss berechnen
            participants_data = []
            total_base_price = 0.0
            total_subsidy = 0.0

            for participant in participants:
                # Alter berechnen
                age = event.start_date.year - participant.birth_date.year
                if (event.start_date.month, event.start_date.day) < (participant.birth_date.month, participant.birth_date.day):
                    age -= 1

                # Basispreis ermitteln
                base_price = PriceCalculator._get_base_price_by_age(
                    age,
                    ruleset.age_groups or []
                )

                # Rollenrabatt berechnen
                discount_percent = role_config.get("discount_percent", 0)
                subsidy_amount = base_price * (discount_percent / 100)

                participants_data.append({
                    "full_name": participant.full_name,
                    "birth_date": participant.birth_date,
                    "age": age,
                    "base_price": base_price,
                    "subsidy_amount": subsidy_amount
                })

                total_base_price += base_price
                total_subsidy += subsidy_amount

            role_subsidies.append({
                "role_id": role.id,
                "role_name": role.name,
                "role_display_name": role.display_name,
                "participants": participants_data,
                "total_base_price": round(total_base_price, 2),
                "total_subsidy": round(total_subsidy, 2)
            })

    # === Kinderrabatt (Familienrabatt) ===
    family_subsidies = None

    if ruleset.family_discount and ruleset.family_discount.get("enabled", False):
        # Alle Kinder (unter 18) mit Familienrabatt laden (ohne manuelle Preisanpassungen)
        participants = db.query(Participant).filter(
            Participant.event_id == event_id,
            Participant.is_active == True,
            Participant.family_id.isnot(None),
            Participant.manual_price_override.is_(None)
        ).all()

        participants_data = []
        total_base_price = 0.0
        total_subsidy = 0.0

        for participant in participants:
            # Alter berechnen
            age = event.start_date.year - participant.birth_date.year
            if (event.start_date.month, event.start_date.day) < (participant.birth_date.month, participant.birth_date.day):
                age -= 1

            # Nur Kinder unter 18
            if age >= 18:
                continue

            # Basispreis ermitteln
            base_price = PriceCalculator._get_base_price_by_age(
                age,
                ruleset.age_groups or []
            )

            # Position in Familie ermitteln
            siblings = db.query(Participant).filter(
                Participant.family_id == participant.family_id,
                Participant.is_active == True,
                Participant.event_id == event_id
            ).order_by(Participant.birth_date).all()

            # Position des Kindes bestimmen (nach Geburtsdatum sortiert)
            child_position = 1
            for idx, sibling in enumerate(siblings, start=1):
                if sibling.id == participant.id:
                    child_position = idx
                    break

            # Familienrabatt berechnen
            family_discount_percent = PriceCalculator._get_family_discount(
                age,
                child_position,
                ruleset.family_discount
            )

            subsidy_amount = base_price * (family_discount_percent / 100)

            # Wenn kein Familienrabatt, überspringe diesen Teilnehmer
            if subsidy_amount == 0:
                continue

            # Familie-Name
            family_name = participant.family.name if participant.family else None

            participants_data.append({
                "full_name": participant.full_name,
                "birth_date": participant.birth_date,
                "age": age,
                "family_name": family_name,
                "base_price": base_price,
                "subsidy_amount": subsidy_amount
            })

            total_base_price += base_price
            total_subsidy += subsidy_amount

        if participants_data:
            family_subsidies = {
                "participants": participants_data,
                "total_base_price": round(total_base_price, 2),
                "total_subsidy": round(total_subsidy, 2)
            }

    logger.info(f"Loaded {len(role_subsidies)} role subsidies and family subsidies: {family_subsidies is not None}")

    return templates.TemplateResponse(
        "cash_status/overview.html",
        {
            "request": request,
            "title": "Zuschüsse",
            "show_subsidies": True,
            "role_subsidies": role_subsidies,
            "family_subsidies": family_subsidies
        }
    )


@router.get("/subsidies/export/pdf")
async def export_subsidy_pdf(
    db: Session = Depends(get_db),
    event_id: int = Depends(get_current_event_id),
    type: str = Query(..., description="Type of subsidy: 'role' or 'family'"),
    role_id: Optional[int] = Query(None, description="Role ID for role-based subsidies")
):
    """
    Exportiert Zuschusslisten als PDF
    """
    logger.info(f"Exporting subsidy PDF for event {event_id}, type={type}, role_id={role_id}")

    # Event laden
    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        logger.error(f"Event {event_id} not found")
        return Response(content="Event nicht gefunden", status_code=404)

    # Aktives Ruleset für das Event finden
    ruleset = db.query(Ruleset).filter(
        Ruleset.event_id == event_id,
        Ruleset.is_active == True,
        Ruleset.valid_from <= event.start_date,
        Ruleset.valid_until >= event.start_date
    ).first()

    if not ruleset:
        logger.error(f"No active ruleset found for event {event_id}")
        return Response(content="Kein aktives Regelwerk gefunden", status_code=404)

    # PDF-Daten basierend auf Typ erstellen
    if type == "role":
        if not role_id:
            return Response(content="Role ID fehlt", status_code=400)

        # Rolle laden
        role = db.query(Role).filter(Role.id == role_id).first()
        if not role:
            return Response(content="Rolle nicht gefunden", status_code=404)

        # Rollenconfig aus Ruleset laden
        role_config = None
        role_name_lower = role.name.lower()
        for key, value in (ruleset.role_discounts or {}).items():
            if key.lower() == role_name_lower:
                role_config = value
                break

        if not role_config:
            return Response(content="Rollenkonfiguration nicht gefunden", status_code=404)

        # Teilnehmer mit dieser Rolle laden (ohne manuelle Preisanpassungen)
        participants = db.query(Participant).filter(
            Participant.event_id == event_id,
            Participant.is_active == True,
            Participant.role_id == role.id,
            Participant.manual_price_override.is_(None)
        ).all()

        # Daten für PDF vorbereiten
        participants_data = []
        total_base_price = 0.0
        total_subsidy = 0.0

        for participant in participants:
            age = event.start_date.year - participant.birth_date.year
            if (event.start_date.month, event.start_date.day) < (participant.birth_date.month, participant.birth_date.day):
                age -= 1

            base_price = PriceCalculator._get_base_price_by_age(age, ruleset.age_groups or [])
            discount_percent = role_config.get("discount_percent", 0)
            subsidy_amount = base_price * (discount_percent / 100)

            participants_data.append({
                "name": participant.full_name,
                "birth_date": participant.birth_date,
                "base_price": base_price,
                "subsidy_amount": subsidy_amount
            })

            total_base_price += base_price
            total_subsidy += subsidy_amount

        # PDF erstellen
        buffer = _create_subsidy_pdf(
            event=event,
            subsidy_type=f"Rollenzuschuss: {role.display_name}",
            participants=participants_data,
            total_subsidy=total_subsidy,
            total_base_price=total_base_price
        )

        filename = f"Zuschussliste_{role.display_name}_{event.name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf"

    elif type == "family":
        # Alle Kinder mit Familienrabatt laden (ohne manuelle Preisanpassungen)
        if not ruleset.family_discount or not ruleset.family_discount.get("enabled", False):
            return Response(content="Familienrabatt nicht aktiviert", status_code=404)

        participants = db.query(Participant).filter(
            Participant.event_id == event_id,
            Participant.is_active == True,
            Participant.family_id.isnot(None),
            Participant.manual_price_override.is_(None)
        ).all()

        participants_data = []
        total_base_price = 0.0
        total_subsidy = 0.0

        for participant in participants:
            age = event.start_date.year - participant.birth_date.year
            if (event.start_date.month, event.start_date.day) < (participant.birth_date.month, participant.birth_date.day):
                age -= 1

            if age >= 18:
                continue

            base_price = PriceCalculator._get_base_price_by_age(age, ruleset.age_groups or [])

            # Position in Familie
            siblings = db.query(Participant).filter(
                Participant.family_id == participant.family_id,
                Participant.is_active == True,
                Participant.event_id == event_id
            ).order_by(Participant.birth_date).all()

            child_position = 1
            for idx, sibling in enumerate(siblings, start=1):
                if sibling.id == participant.id:
                    child_position = idx
                    break

            family_discount_percent = PriceCalculator._get_family_discount(
                age, child_position, ruleset.family_discount
            )

            subsidy_amount = base_price * (family_discount_percent / 100)

            if subsidy_amount == 0:
                continue

            participants_data.append({
                "name": participant.full_name,
                "birth_date": participant.birth_date,
                "base_price": base_price,
                "subsidy_amount": subsidy_amount
            })

            total_base_price += base_price
            total_subsidy += subsidy_amount

        # PDF erstellen
        buffer = _create_subsidy_pdf(
            event=event,
            subsidy_type="Kinderrabatt (MGB-Zuschuss)",
            participants=participants_data,
            total_subsidy=total_subsidy,
            total_base_price=total_base_price
        )

        filename = f"Zuschussliste_Kinderrabatt_{event.name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf"

    else:
        return Response(content="Ungültiger Typ", status_code=400)

    logger.info(f"PDF export completed: {filename}")

    return Response(
        content=buffer.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/subsidies/export/excel")
async def export_subsidy_excel(
    db: Session = Depends(get_db),
    event_id: int = Depends(get_current_event_id),
    type: str = Query(..., description="Type of subsidy: 'role' or 'family'"),
    role_id: Optional[int] = Query(None, description="Role ID for role-based subsidies")
):
    """
    Exportiert Zuschusslisten als Excel-Datei
    """
    logger.info(f"Exporting subsidy Excel for event {event_id}, type={type}, role_id={role_id}")

    # Event laden
    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        logger.error(f"Event {event_id} not found")
        return Response(content="Event nicht gefunden", status_code=404)

    # Aktives Ruleset für das Event finden
    ruleset = db.query(Ruleset).filter(
        Ruleset.event_id == event_id,
        Ruleset.is_active == True,
        Ruleset.valid_from <= event.start_date,
        Ruleset.valid_until >= event.start_date
    ).first()

    if not ruleset:
        logger.error(f"No active ruleset found for event {event_id}")
        return Response(content="Kein aktives Regelwerk gefunden", status_code=404)

    # Excel-Daten basierend auf Typ erstellen
    if type == "role":
        if not role_id:
            return Response(content="Role ID fehlt", status_code=400)

        # Rolle laden
        role = db.query(Role).filter(Role.id == role_id).first()
        if not role:
            return Response(content="Rolle nicht gefunden", status_code=404)

        # Rollenconfig aus Ruleset laden
        role_config = None
        role_name_lower = role.name.lower()
        for key, value in (ruleset.role_discounts or {}).items():
            if key.lower() == role_name_lower:
                role_config = value
                break

        if not role_config:
            return Response(content="Rollenkonfiguration nicht gefunden", status_code=404)

        # Teilnehmer mit dieser Rolle laden (ohne manuelle Preisanpassungen)
        participants = db.query(Participant).filter(
            Participant.event_id == event_id,
            Participant.is_active == True,
            Participant.role_id == role.id,
            Participant.manual_price_override.is_(None)
        ).all()

        # Daten für Excel vorbereiten
        participants_data = []
        total_base_price = 0.0
        total_subsidy = 0.0

        for participant in participants:
            age = event.start_date.year - participant.birth_date.year
            if (event.start_date.month, event.start_date.day) < (participant.birth_date.month, participant.birth_date.day):
                age -= 1

            base_price = PriceCalculator._get_base_price_by_age(age, ruleset.age_groups or [])
            discount_percent = role_config.get("discount_percent", 0)
            subsidy_amount = base_price * (discount_percent / 100)

            participants_data.append({
                "name": participant.full_name,
                "birth_date": participant.birth_date,
                "base_price": base_price,
                "subsidy_amount": subsidy_amount
            })

            total_base_price += base_price
            total_subsidy += subsidy_amount

        # Excel erstellen
        buffer = _create_subsidy_excel(
            event=event,
            subsidy_type=f"Rollenzuschuss: {role.display_name}",
            participants=participants_data,
            total_subsidy=total_subsidy,
            total_base_price=total_base_price
        )

        filename = f"Zuschussliste_{role.display_name}_{event.name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"

    elif type == "family":
        # Alle Kinder mit Familienrabatt laden (ohne manuelle Preisanpassungen)
        if not ruleset.family_discount or not ruleset.family_discount.get("enabled", False):
            return Response(content="Familienrabatt nicht aktiviert", status_code=404)

        participants = db.query(Participant).filter(
            Participant.event_id == event_id,
            Participant.is_active == True,
            Participant.family_id.isnot(None),
            Participant.manual_price_override.is_(None)
        ).all()

        participants_data = []
        total_base_price = 0.0
        total_subsidy = 0.0

        for participant in participants:
            age = event.start_date.year - participant.birth_date.year
            if (event.start_date.month, event.start_date.day) < (participant.birth_date.month, participant.birth_date.day):
                age -= 1

            if age >= 18:
                continue

            base_price = PriceCalculator._get_base_price_by_age(age, ruleset.age_groups or [])

            # Position in Familie
            siblings = db.query(Participant).filter(
                Participant.family_id == participant.family_id,
                Participant.is_active == True,
                Participant.event_id == event_id
            ).order_by(Participant.birth_date).all()

            child_position = 1
            for idx, sibling in enumerate(siblings, start=1):
                if sibling.id == participant.id:
                    child_position = idx
                    break

            family_discount_percent = PriceCalculator._get_family_discount(
                age, child_position, ruleset.family_discount
            )

            subsidy_amount = base_price * (family_discount_percent / 100)

            if subsidy_amount == 0:
                continue

            participants_data.append({
                "name": participant.full_name,
                "birth_date": participant.birth_date,
                "base_price": base_price,
                "subsidy_amount": subsidy_amount
            })

            total_base_price += base_price
            total_subsidy += subsidy_amount

        # Excel erstellen
        buffer = _create_subsidy_excel(
            event=event,
            subsidy_type="Kinderrabatt (MGB-Zuschuss)",
            participants=participants_data,
            total_subsidy=total_subsidy,
            total_base_price=total_base_price
        )

        filename = f"Zuschussliste_Kinderrabatt_{event.name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"

    else:
        return Response(content="Ungültiger Typ", status_code=400)

    logger.info(f"Excel export completed: {filename}")

    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


def _create_subsidy_excel(
    event: Event,
    subsidy_type: str,
    participants: list,
    total_subsidy: float,
    total_base_price: float
) -> BytesIO:
    """
    Erstellt eine Excel-Datei für eine Zuschussliste

    Args:
        event: Das Event
        subsidy_type: Art des Zuschusses (z.B. "Rollenzuschuss: Betreuer")
        participants: Liste mit Teilnehmerdaten
        total_subsidy: Gesamtsumme der Zuschüsse
        total_base_price: Gesamtsumme der Basispreise

    Returns:
        BytesIO mit Excel-Inhalt
    """
    # Workbook erstellen
    wb, ws = ExcelService.create_workbook(f"{subsidy_type}")

    # Header-Informationen
    today = date.today()
    ws['A1'] = event.name
    ws['A1'].font = Font(size=16, bold=True, color='1e40af')

    ws['E1'] = f"Beantragungsdatum: {today.strftime('%d.%m.%Y')}"
    ws['E1'].font = Font(size=10)
    ws['E1'].alignment = Alignment(horizontal='right')

    ws['A2'] = f"Zeitraum: {event.start_date.strftime('%d.%m.%Y')} - {event.end_date.strftime('%d.%m.%Y')}"
    ws['A2'].font = Font(size=10)

    ws['A3'] = f"Art des Zuschusses: {subsidy_type}"
    ws['A3'].font = Font(size=12, bold=True, color='1e40af')

    # Tabellen-Header (Zeile 5)
    headers = ["Name", "Geburtsdatum", "Regulärer Preis (€)", "Zuschuss (€)"]
    column_widths = {1: 30, 2: 20, 3: 20, 4: 20}

    ExcelService.apply_header_row(ws, headers, column_widths, start_row=5)

    # Daten schreiben
    row_num = 6
    for participant in participants:
        ws.cell(row=row_num, column=1, value=participant["name"])
        ws.cell(row=row_num, column=2, value=participant["birth_date"].strftime('%d.%m.%Y'))
        ws.cell(row=row_num, column=3, value=participant["base_price"])
        ws.cell(row=row_num, column=4, value=participant["subsidy_amount"])

        # Farbmarkierung für Zuschuss
        ExcelService.apply_color_by_value(ws, row_num, 4, participant["subsidy_amount"])

        row_num += 1

    # Summenzeile
    row_num += 1
    summary_values = {
        3: total_base_price,
        4: total_subsidy
    }
    ExcelService.apply_summary_row(ws, row_num, summary_values)

    ws.cell(row=row_num, column=1, value="Gesamt")
    ws.cell(row=row_num, column=1).font = Font(bold=True)

    # Unterschriftenfeld
    row_num += 3
    ws.cell(row=row_num, column=1, value="Unterschrift:")
    ws.cell(row=row_num, column=1).font = Font(bold=True)

    row_num += 1
    ws.cell(row=row_num, column=1, value="_" * 50)
    ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')

    # Datei als BytesIO zurückgeben
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output


def _create_subsidy_pdf(
    event: Event,
    subsidy_type: str,
    participants: list,
    total_subsidy: float,
    total_base_price: float
) -> BytesIO:
    """
    Erstellt ein PDF für eine Zuschussliste

    Args:
        event: Das Event
        subsidy_type: Art des Zuschusses (z.B. "Rollenzuschuss: Betreuer")
        participants: Liste mit Teilnehmerdaten
        total_subsidy: Gesamtsumme der Zuschüsse
        total_base_price: Gesamtsumme der Basispreise

    Returns:
        BytesIO mit PDF-Inhalt
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        topMargin=2*cm,
        bottomMargin=2*cm,
        leftMargin=2*cm,
        rightMargin=2*cm
    )
    story = []

    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=10,
        alignment=0  # Left
    )
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=11,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=5,
    )
    normal_style = styles['Normal']
    small_style = ParagraphStyle(
        'Small',
        parent=styles['Normal'],
        fontSize=9,
    )

    # === Header ===
    # Titel und Beantragungsdatum
    today = date.today()
    header_data = [
        [Paragraph(event.name, title_style), Paragraph(f"Beantragungsdatum: {today.strftime('%d.%m.%Y')}", small_style)]
    ]
    header_table = Table(header_data, colWidths=[12*cm, 5*cm])
    header_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 0.3*cm))

    # Zeitraum der Freizeit
    period_text = f"Zeitraum: {event.start_date.strftime('%d.%m.%Y')} - {event.end_date.strftime('%d.%m.%Y')}"
    story.append(Paragraph(period_text, small_style))
    story.append(Spacer(1, 0.5*cm))

    # Art des Zuschusses
    story.append(Paragraph(f"<b>Art des Zuschusses:</b> {subsidy_type}", heading_style))
    story.append(Spacer(1, 0.5*cm))

    # === Teilnehmer-Tabelle ===
    table_data = [
        ["Name", "Geburtsdatum", "Regulärer Preis", "Zuschuss"]
    ]

    for participant in participants:
        table_data.append([
            participant["name"],
            participant["birth_date"].strftime('%d.%m.%Y'),
            f"{participant['base_price']:.2f} €",
            f"{participant['subsidy_amount']:.2f} €"
        ])

    # Tabelle erstellen
    participants_table = Table(table_data, colWidths=[6*cm, 3.5*cm, 3.5*cm, 3.5*cm])

    # Tabellen-Style
    table_style = [
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (1, 0), (-1, 0), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),

        # Daten
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),
        ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),

        # Gitter
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),

        # Alternierende Zeilen
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
    ]

    participants_table.setStyle(TableStyle(table_style))
    story.append(participants_table)
    story.append(Spacer(1, 0.8*cm))

    # === Summe ===
    summary_data = [
        ["Gesamtsumme der Zuschüsse:", f"{total_subsidy:.2f} €"]
    ]
    summary_table = Table(summary_data, colWidths=[10*cm, 6.5*cm])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#E8F4F8')),
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#1e40af')),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ('TEXTCOLOR', (1, 0), (1, 0), colors.HexColor('#00B050')),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 1.5*cm))

    # === Unterschriftenfeld ===
    story.append(Paragraph("<b>Unterschrift:</b>", normal_style))
    story.append(Spacer(1, 0.3*cm))

    # Linie für Unterschrift
    signature_line = Table([["_" * 80]], colWidths=[16.5*cm])
    signature_line.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.grey),
    ]))
    story.append(signature_line)

    # PDF generieren
    doc.build(story)
    buffer.seek(0)

    return buffer
